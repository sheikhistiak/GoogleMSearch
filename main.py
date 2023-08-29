import openpyxl
from GoogleBMFunc import extract_search_results
from LoadXL_DataFunc import load_column_data_from_excel

excel_file_path = 'Search_meaning.xlsx'
data = load_column_data_from_excel(excel_file_path)

start_row = 2
# Print each item in the data list
for item in data:
    print(item)
    result = extract_search_results(item)
    # print(result)

    search_result = result

    if "Meaning in Bangla:" in search_result:
        # Extract the English meaning
        english_meaning = search_result.split("\n")[2].strip()
        print("English Meaning:", english_meaning)

        # Extract the Bangla translation
        bangla_translation = search_result.split("Meaning in Bangla: \n")[1].strip()
        print("Bangla Translation:", bangla_translation)
    else:
        print("Meaning in Bangla not found in the search result.")
        bangla_translation = "Meaning in Bangla not found in the search result."
        pass

    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active  # Assume you want to work with the active sheet
    # Write the values into the worksheet
    sheet.cell(row=start_row, column=4, value=result)
    sheet.cell(row=start_row, column=5, value=bangla_translation)
    start_row += 1

    # Save the changes to the Excel file
    workbook.save(excel_file_path)
